"""
One-shot enrichment: geocode + fill levels + Facebook/website source
into shool.xlsx for the 10 new schools (กระบี่ rows 2-9, ระนอง rows 30-36).

Levels are inferred from the existing rich-text fields (col 14-19) which
the user maintains; Facebook URLs come from the verified web searches.
"""
import sys, io, json, time, urllib.parse, urllib.request
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

CHECK = "✓"
DASH = "—"

# School-level enrichment, keyed by the "primary" row of each unique school.
# Duplicates (rows 4, 31-33, 35) reuse the same key.
ENRICHMENT = {
    2:  {  # โรงเรียนเทศบาลอ่าวลึกใต้  — ต.อ่าวลึกใต้ อ.อ่าวลึก จ.กระบี่
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": None,                                  # not found
        "website":  None,
        "lat": 8.3760, "lng": 98.7310,                     # tambon center (approx)
    },
    3:  {  # โรงเรียนเทศบาลคลองท่อมใต้  — ต.คลองท่อมใต้ อ.คลองท่อม จ.กระบี่
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/tbktschool/",
        "website":  "https://tbktschool.ac.th",
        "lat": 7.9170, "lng": 99.1580,
    },
    5:  {  # โรงเรียนเทศบาล 4 มหาราช  — ต.กระบี่ใหญ่ อ.เมืองกระบี่ จ.กระบี่
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/MaharachMunicipalSchool4/",
        "website":  None,
        "lat": 8.0720, "lng": 98.9270,
    },
    6:  {  # โรงเรียนเทศบาล 3 ท่าแดง  — 290 ถ.มหาราช ต.ปากน้ำ อ.เมืองกระบี่
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/100067055134888/",
        "website":  None,
        "lat": 8.0610, "lng": 98.9080,
    },
    7:  {  # โรงเรียนอนุบาลบางเท่าแม่  — ต.เขาต่อ อ.ปลายพระยา จ.กระบี่ (อนุบาลล้วน)
        "level_range": "อ.1 - อ.3",
        "k": True, "p": False, "ls": False, "us": False,
        "facebook": "https://www.facebook.com/xnubalbangtheam/",
        "website":  "https://anubanbangtaomae.ac.th",
        "lat": 8.4720, "lng": 98.8210,
    },
    8:  {  # โรงเรียนบ้านช้างตาย  — ต.เขาดิน อ.เขาพนม จ.กระบี่
        "level_range": "อ.1 - ป.6",
        "k": True, "p": True, "ls": False, "us": False,
        "facebook": "https://www.facebook.com/100095514283778/",
        "website":  None,
        "lat": 8.1760, "lng": 99.1150,
    },
    9:  {  # โรงเรียนบ้านช่องพลี  — ต.อ่าวนาง อ.เมืองกระบี่ จ.กระบี่ (มีฝ่ายมัธยม)
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/chongphlischool/",
        "website":  "https://chongplee.ac.th",
        "lat": 8.0390, "lng": 98.8210,
    },
    30: {  # โรงเรียนเทศบาลวัดอุปนันทาราม  — 134 ถ.ท่าเมือง ต.เขานิเวศน์ อ.เมืองระนอง
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/100057330825404/",
        "website":  "https://uppanan-ranong.ac.th",
        "lat": 9.9650, "lng": 98.6400,
    },
    34: {  # โรงเรียนเทศบาลบ้านเขานิเวศน์  — 4 ซ.ลุวัง ต.เขานิเวศน์ อ.เมืองระนอง
        "level_range": "อ.1 - ม.3",
        "k": True, "p": True, "ls": True, "us": False,
        "facebook": "https://www.facebook.com/361086113902996/",
        "website":  "https://khaoniwat.ac.th",
        "lat": 9.9620, "lng": 98.6360,
    },
    36: {  # โรงเรียนบ้านในวง  — ต.ในวงเหนือ อ.ละอุ่น จ.ระนอง (ประถมล้วน)
        "level_range": "ป.1 - ป.6",
        "k": False, "p": True, "ls": False, "us": False,
        "facebook": "https://www.facebook.com/naiwong.school/",
        "website":  None,
        "lat": 9.7780, "lng": 98.6960,
    },
}

# Map every target row to its enrichment key (handles duplicates).
ROW_TO_KEY = {
    2: 2, 3: 3, 4: 3, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9,
    30: 30, 31: 30, 32: 30, 33: 30, 34: 34, 35: 34, 36: 36,
}


def geocode(query):
    """Hit Nominatim once. Free, ~1 req/sec, no API key."""
    url = (
        "https://nominatim.openstreetmap.org/search?"
        f"q={urllib.parse.quote(query)}&format=json&limit=1&countrycodes=th"
    )
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "phuket-lao-schools/1.0 enrich_xlsx"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode("utf-8"))
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"    geocode error: {e}")
    return None, None


def derive_source_text(fb_url, website_url):
    """Build the text of the source cell from the URLs we have."""
    parts = []
    if fb_url:
        slug = (fb_url
                .replace("https://www.facebook.com/", "facebook.com/")
                .replace("https://facebook.com/", "facebook.com/")
                .rstrip("/"))
        parts.append(slug)
    if website_url:
        slug = (website_url
                .replace("https://", "")
                .replace("http://", "")
                .rstrip("/"))
        parts.append(slug)
    return " + ".join(parts) if parts else None


def main():
    wb = load_workbook("shool.xlsx")
    ws = wb["Phuket"]

    for row_idx in sorted(ROW_TO_KEY):
        e = ENRICHMENT[ROW_TO_KEY[row_idx]]
        name = ws.cell(row=row_idx, column=4).value
        print(f"\nrow {row_idx}: {name}")

        # ── Map cell — use hard-coded tambon-center coords ──
        lat, lng = e["lat"], e["lng"]
        map_url = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
        cell = ws.cell(row=row_idx, column=6)
        cell.value = "เปิดแผนที่"
        cell.hyperlink = map_url
        print(f"  → map: {lat}, {lng}")

        # ── Levels ──
        ws.cell(row=row_idx, column=7).value  = e["level_range"]
        ws.cell(row=row_idx, column=8).value  = CHECK if e["k"]  else DASH
        ws.cell(row=row_idx, column=9).value  = CHECK if e["p"]  else DASH
        ws.cell(row=row_idx, column=10).value = CHECK if e["ls"] else DASH
        ws.cell(row=row_idx, column=11).value = CHECK if e["us"] else DASH

        # ── Source cell (FB + website, with hyperlink to FB if present) ──
        text = derive_source_text(e.get("facebook"), e.get("website"))
        if text:
            cell = ws.cell(row=row_idx, column=12)
            cell.value = text
            cell.hyperlink = e.get("facebook") or e.get("website")
            print(f"  → src:  {text}")
        else:
            print(f"  → src:  (none)")

    wb.save("shool.xlsx")
    print("\nSaved shool.xlsx")


if __name__ == "__main__":
    main()
