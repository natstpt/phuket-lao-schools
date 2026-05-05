"""
Read shool.xlsx, filter Phuket rows, merge duplicates, write data.js.

Same-school criterion: school name + address.
- positions[] collects every เลขที่ตำแหน่ง for the merged school
- vacancy is the sum of อัตราว่าง across merged rows
- levels are the union (any ✓ across merged rows)
- google_map keeps the first hyperlink seen
- sources are parsed into [{url, label}] from text like "domain.com + facebook.com/page",
  also picking up the raw cell hyperlink if present
"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

CHECK = "✓"

LEVEL_KEYS = ["kindergarten", "primary", "lower_secondary", "upper_secondary"]

# Filter to schools in these southern provinces.
SOUTHERN_PROVINCES = {"ภูเก็ต", "กระบี่", "ระนอง"}

# Authoritative Facebook URLs by school name. If a school is listed here,
# any FB links pulled from the xlsx are dropped and only these are used —
# avoids duplicate-FB issues when Thai-character URL paths don't dedup.
#
# Value can be:
#   • a single URL string                        → labeled "Facebook"
#   • a list of strings/tuples                    → multiple FB pages, in order
#   • a tuple (url, label) inside the list        → custom label
MANUAL_FACEBOOK = {
    "โรงเรียนเทศบาลอ่าวลึกใต้":               "https://www.facebook.com/people/โรงเรียนเทศบาลอ่าวลึกใต้-จังหวัดกระบี่/100063505122666/",
    "โรงเรียนเทศบาลพิบูลสวัสดี":              "https://www.facebook.com/184447268433225/",
    "โรงเรียนเทศบาลปลูกปัญญาฯ":               "https://www.facebook.com/plukpanyaschool/",
    "โรงเรียนเทศบาลบ้านสามกองฯ":              "https://www.facebook.com/SamkongSchool/",
    "โรงเรียน อบจ.บ้านไม้เรียบ (ตันติโกวิทบำรุง)": "https://www.facebook.com/banmaireab/",
    "โรงเรียน อบจ.บ้านนาบอน":                 "https://www.facebook.com/100063651889627/",
    "โรงเรียน อบจ.บ้านตลาดเหนือ (วันครู 2502)":  "https://www.facebook.com/wankroo2502/",
    "โรงเรียนบ้านช่องพลี": [
        "https://www.facebook.com/cp.ac.th/",
        ("https://www.facebook.com/p/โรงเรียนบ้านช่องพลี-ฝ่ายมัธยม-100065335666012/",
         "Facebook (ฝ่ายมัธยม)"),
    ],
}

# Domains to drop from the source list (e.g. unsafe / broken sites).
EXCLUDE_DOMAINS = {"nabon.ac.th"}

# Hardcoded fallback coordinates (tambon-center, ~hundreds of meters
# accuracy). Used when the xlsx Google Map cell has a cid=... share URL
# we can't extract lat/lng from. The clickable URL still points at the
# user's preferred (accurate) Google Maps target — only the Leaflet pin
# uses these fallback coords.
MANUAL_COORDS = {
    "โรงเรียนเทศบาลอ่าวลึกใต้":      (8.3760, 98.7310),
    "โรงเรียนเทศบาลคลองท่อมใต้":     (7.9170, 99.1580),
    "โรงเรียนเทศบาล 4 มหาราช":        (8.0720, 98.9270),
    "โรงเรียนเทศบาล 3 ท่าแดง":        (8.0610, 98.9080),
    "โรงเรียนอนุบาลบางเท่าแม่":      (8.4720, 98.8210),
    "โรงเรียนบ้านช้างตาย":            (8.1760, 99.1150),
    "โรงเรียนบ้านช่องพลี":            (8.0390, 98.8210),
    "โรงเรียนเทศบาลวัดอุปนันทาราม":  (9.9650, 98.6400),
    "โรงเรียนเทศบาลบ้านเขานิเวศน์":  (9.9620, 98.6360),
    "โรงเรียนบ้านในวง":               (9.7780, 98.6960),
}


def is_check(v):
    return isinstance(v, str) and CHECK in v


def extract_province(address):
    """Pull the province name out of a Thai address line like '...จ.กระบี่ 81000'."""
    if not address:
        return None
    m = re.search(r"จ\.(\S+)", str(address))
    if not m:
        return None
    prov = m.group(1)
    # Strip trailing digits / punctuation that may have stuck on
    prov = re.sub(r"[\d\s].*$", "", prov)
    return prov.strip() or None


def clean_text(v):
    """Strip whitespace; treat empty / lone em-dash as no value."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "—" or s == "-":
        return None
    return s


def normalize_url(u):
    if not u:
        return None
    u = u.strip()
    if not u:
        return None
    if not re.match(r"^https?://", u, re.I):
        u = "https://" + u
    return u


def url_key(u):
    """Canonical key for dedup: strip protocol, www., trailing slash, lowercase."""
    if not u:
        return ""
    k = re.sub(r"^https?://", "", u, flags=re.I)
    k = re.sub(r"^www\.", "", k, flags=re.I)
    return k.rstrip("/").lower()


def domain_of(u):
    """Just the host portion (no path), lowercased."""
    return url_key(u).split("/")[0]


def ensure_coords(school):
    """
    If the xlsx Google Map URL didn't yield lat/lng (e.g. it's a
    `?cid=...` share link), fall back to MANUAL_COORDS so the Leaflet
    pin still appears. The clickable URL is preserved so users still
    land on the user's chosen Google Maps target.
    """
    name = school["name"]
    if name not in MANUAL_COORDS:
        return
    gm = school.get("google_map")
    lat, lng = MANUAL_COORDS[name]
    if not gm:
        school["google_map"] = {
            "url": f"https://www.google.com/maps/search/?api=1&query={lat},{lng}",
            "lat": lat,
            "lng": lng,
        }
        return
    if gm.get("lat") is None or gm.get("lng") is None:
        gm["lat"] = lat
        gm["lng"] = lng


def _normalize_fb_entries(raw):
    """Coerce a MANUAL_FACEBOOK value into a list of {url, label} dicts.
    Accepts a plain string, a list of strings, or a list with (url, label)
    tuples for custom labels (e.g. "Facebook (ฝ่ายมัธยม)")."""
    items = raw if isinstance(raw, list) else [raw]
    out = []
    for it in items:
        if isinstance(it, tuple):
            out.append({"url": it[0], "label": it[1]})
        else:
            out.append({"url": it, "label": "Facebook"})
    return out


def fixup_sources(school):
    """
    Post-process a school's sources list:
      1. Drop any URL whose domain is in EXCLUDE_DOMAINS.
      2. If MANUAL_FACEBOOK has an entry, those FB URLs are the single
         source of truth — strip every other facebook.com link, then
         prepend the manual ones in their declared order. Avoids
         duplicate FB entries that slip past url_key dedup.
      3. Otherwise just move whatever Facebook URL the xlsx provided to
         the top of the list.
    """
    sources = [s for s in school["sources"] if domain_of(s["url"]) not in EXCLUDE_DOMAINS]

    if school["name"] in MANUAL_FACEBOOK:
        sources = [s for s in sources if "facebook.com" not in s["url"].lower()]
        fb_entries = _normalize_fb_entries(MANUAL_FACEBOOK[school["name"]])
        sources = fb_entries + sources
    else:
        fb_idx = next((i for i, s in enumerate(sources) if "facebook.com" in s["url"].lower()), -1)
        if fb_idx > 0:
            sources.insert(0, sources.pop(fb_idx))

    school["sources"] = sources


def label_for(url):
    """Build a human-readable label for a URL."""
    if not url:
        return ""
    bare = re.sub(r"^https?://", "", url, flags=re.I).rstrip("/")
    if "facebook.com" in bare.lower():
        return "Facebook"
    # Long: keep just the domain
    if len(bare) > 30:
        domain = bare.split("/")[0]
        return domain
    return bare


def parse_sources(text, hyperlink):
    """
    Parse a cell like 'facebook.com/foo + example.com/bar' into [{url, label}, ...].
    The cell typically carries a hyperlink to one of the listed URLs; we still
    surface every textual entry so the user can find them all.
    """
    items = []
    seen = set()

    def push(raw):
        url = normalize_url(raw)
        if not url:
            return
        k = url_key(url)
        if k in seen:
            return
        seen.add(k)
        items.append({"url": url, "label": label_for(url)})

    if text:
        for chunk in re.split(r"\s*\+\s*", str(text)):
            chunk = chunk.strip()
            if chunk:
                push(chunk)

    if hyperlink:
        push(hyperlink)

    return items


def parse_map(map_url):
    """Return {url, lat, lng} from a Google Maps URL with ?query=lat,lng."""
    if not map_url:
        return None
    m = re.search(r"query=([\-\d.]+),([\-\d.]+)", map_url)
    coords = None
    if m:
        coords = {"lat": float(m.group(1)), "lng": float(m.group(2))}
    return {"url": map_url, **(coords or {})}


def main():
    wb = load_workbook("shool.xlsx", data_only=True)
    ws = wb["Phuket"]

    schools = {}
    order = []

    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=c) for c in range(1, 24)]
        (seq, org, pos_no, name, address, _map_text, level_range,
         k, p, ls, us, src_text, vac,
         student_count, staff_count, special_programs,
         awards, highlights, extra_sources,
         nearby_places, dorms_condos, rental_houses, landmarks) = (c.value for c in cells)

        province = extract_province(address)
        if not address or province not in SOUTHERN_PROVINCES:
            continue
        if not name:
            continue

        name = str(name).strip()
        address = str(address).strip()
        key = (name, address)

        map_cell = cells[5]
        src_cell = cells[11]
        map_link = map_cell.hyperlink.target if map_cell.hyperlink else None
        src_link = src_cell.hyperlink.target if src_cell.hyperlink else None

        if key not in schools:
            schools[key] = {
                "id": len(schools) + 1,
                "name": name,
                "org": str(org).strip() if org else "",
                "province": province,
                "address": address,
                "levels": str(level_range).strip() if level_range else "",
                "kindergarten": is_check(k),
                "primary": is_check(p),
                "lower_secondary": is_check(ls),
                "upper_secondary": is_check(us),
                "google_map": parse_map(map_link),
                "sources": parse_sources(src_text, src_link),
                "vacancy": 0,
                "positions": [],
                "seq_numbers": [],
                # Optional rich detail fields — None if absent/blank
                "student_count": None,
                "staff_count": None,
                "special_programs": None,
                "awards": None,
                "highlights": None,
                "extra_sources": None,
                # Area / housing fields
                "nearby_places": None,
                "dorms_condos": None,
                "rental_houses": None,
                "landmarks": None,
            }
            order.append(key)

        s = schools[key]

        # Union levels (any row that marks a level keeps it on)
        s["kindergarten"] = s["kindergarten"] or is_check(k)
        s["primary"] = s["primary"] or is_check(p)
        s["lower_secondary"] = s["lower_secondary"] or is_check(ls)
        s["upper_secondary"] = s["upper_secondary"] or is_check(us)

        # Prefer the first non-empty range string we saw
        if not s["levels"] and level_range:
            s["levels"] = str(level_range).strip()

        # Append unique position numbers
        if pos_no:
            pos_str = str(pos_no).strip()
            if pos_str and pos_str not in s["positions"]:
                s["positions"].append(pos_str)

        # Capture original xlsx ลำดับ — show as "1, 2" in the new column
        if seq is not None:
            try:
                seq_int = int(seq) if isinstance(seq, (int, float)) else int(str(seq).strip())
                if seq_int not in s["seq_numbers"]:
                    s["seq_numbers"].append(seq_int)
            except (ValueError, TypeError):
                pass

        # Rich detail fields — keep first non-empty value seen across merged rows
        for field, raw in [
            ("student_count", student_count),
            ("staff_count", staff_count),
            ("special_programs", special_programs),
            ("awards", awards),
            ("highlights", highlights),
            ("extra_sources", extra_sources),
            ("nearby_places", nearby_places),
            ("dorms_condos", dorms_condos),
            ("rental_houses", rental_houses),
            ("landmarks", landmarks),
        ]:
            if not s[field]:
                cleaned = clean_text(raw)
                if cleaned:
                    s[field] = cleaned

        # Sum vacancy
        if isinstance(vac, (int, float)):
            s["vacancy"] += int(vac)
        elif isinstance(vac, str) and vac.strip().isdigit():
            s["vacancy"] += int(vac.strip())

        # Merge sources picked up across duplicate rows
        existing_keys = {url_key(x["url"]) for x in s["sources"]}
        for item in parse_sources(src_text, src_link):
            if url_key(item["url"]) not in existing_keys:
                s["sources"].append(item)
                existing_keys.add(url_key(item["url"]))

        # Use the first map link we saw; keep it if already set
        if not s["google_map"] and map_link:
            s["google_map"] = parse_map(map_link)

    # Build the final list and post-process sources + coords.
    final = []
    for key in order:
        rec = schools[key]
        fixup_sources(rec)
        ensure_coords(rec)
        final.append(rec)

    # Sort by the smallest original xlsx ลำดับ so that cards/rows render
    # left-to-right, top-to-bottom in numerical order regardless of
    # how the xlsx rows happened to be grouped.
    final.sort(key=lambda s: min(s["seq_numbers"]) if s["seq_numbers"] else 9_999)
    for i, rec in enumerate(final, 1):
        rec["id"] = i

    payload = json.dumps(final, ensure_ascii=False, indent=2)

    with open("data.js", "w", encoding="utf-8") as f:
        f.write("// Auto-generated from shool.xlsx — do not edit by hand.\n")
        f.write(f"// Source rows from sheet 'Phuket', filtered where address contains 'ภูเก็ต'.\n")
        f.write(f"const SCHOOLS = {payload};\n")

    # Console summary
    print(f"Wrote data.js — {len(final)} unique schools")
    print(f"Total vacancy: {sum(s['vacancy'] for s in final)}")
    print(f"Orgs: {sorted({s['org'] for s in final})}")
    for s in final:
        print(f"  #{s['id']:2d} {s['name']} — {len(s['positions'])} positions, vac={s['vacancy']}")


if __name__ == "__main__":
    main()
